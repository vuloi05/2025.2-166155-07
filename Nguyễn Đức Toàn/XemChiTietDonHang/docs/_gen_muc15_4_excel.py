#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate Excel for SRS section 15.4 — Kiểm thử Use Case (UC006)."""
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / ".excel_tools"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

OUT = Path(__file__).resolve().parent / "15.4-KiemThu-UseCase-Scenario-Testing.xlsx"

BOLD = Font(bold=True, size=11)
NORMAL = Font(size=11)
TITLE = Font(bold=True, size=12)
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_table(ws, r1, c1, r2, c2, header_row=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP
            if header_row and r == header_row:
                cell.font = BOLD
                cell.alignment = CENTER
            elif r > (header_row or r2):
                cell.font = NORMAL


SUMMARY_ROWS = [
    (
        "UC_TC01",
        "Xem chi tiết đơn hàng thành công (Luồng chính bước 1–9)",
        'Người dùng đã đăng nhập; hệ thống có dữ liệu đơn hàng (13 đơn); '
        'ứng dụng mở tại Trang chủ (M0); tồn tại đơn DH-2026-003 trạng thái "Đã xử lý".',
        'M0 → click ô "Xem chi tiết đơn hàng" → M1 Dashboard → "Xem danh sách đơn hàng" → M2. '
        'Click "Xem" đơn DH-2026-003 → mở modal Chi tiết (M3): hiển thị thông tin chung '
        '(mã đơn, trạng thái, Site, PT vận chuyển, ngày…) và bảng mặt hàng (MH004). '
        "Không có dữ liệu nào bị thay đổi.",
        "PASSED",
    ),
    (
        "UC_TC02",
        "Tìm kiếm không thấy đơn hàng (Luồng thay thế 5a → 6a)",
        "Đang ở màn hình danh sách đơn hàng (M2).",
        'Nhập từ khóa không tồn tại (ví dụ DH-9999) → nhấn "Tìm kiếm" → hiển thị trạng thái rỗng '
        '"Không tìm thấy đơn hàng nào" (M4). Nhấn "Làm mới" → hiển thị lại toàn bộ danh sách trên M2.',
        "PASSED",
    ),
    (
        "UC_TC03",
        "Xem chi tiết đơn hàng đã bị hủy (Luồng thay thế 7a → 7b)",
        'Trong danh sách (M2) có đơn DH-2026-008 trạng thái "Đã hủy".',
        'Click "Xem" đơn DH-2026-008 → hệ thống hiện hộp thoại cảnh báo (M5): '
        '"Đơn hàng đã bị hủy, không thể xem chi tiết". '
        'Nhấn "Xác nhận & Quay lại" → quay về M2; không mở modal chi tiết (M3).',
        "PASSED",
    ),
    (
        "UC_TC04",
        "Lọc theo trạng thái rồi xem chi tiết (Luồng chính bước 3–8)",
        "Đang ở màn hình danh sách đơn hàng (M2).",
        'Chọn bộ lọc Trạng thái = "Đã xử lý" → nhấn "Tìm kiếm" → bảng chỉ còn các đơn phù hợp '
        "(DH-2026-003, DH-2026-011). Click \"Xem\" một đơn → mở đúng modal chi tiết (M3).",
        "PASSED",
    ),
    (
        "UC_TC05",
        "Quay lại danh sách sau khi xem chi tiết (Luồng chính bước 9)",
        "Đang ở modal chi tiết đơn hàng (M3) của một đơn hợp lệ.",
        'Nhấn "← Quay lại danh sách" (hoặc nút "✕") → đóng modal, trở về màn hình danh sách (M2). '
        "Không có dữ liệu nào bị thay đổi.",
        "PASSED",
    ),
]

DETAIL_CASES = [
    {
        "id": "UC_TC01",
        "title": "Xem chi tiết đơn hàng thành công",
        "pre": "Người dùng đã đăng nhập; hệ thống có 13 đơn hàng mock; tồn tại DH-2026-003 (Đã xử lý).",
        "steps": [
            ("1", 'Khởi chạy ứng dụng', "Hiển thị Trang chủ hệ thống (M0) — lưới 6 ô chức năng"),
            ("2", 'Click ô "Xem chi tiết đơn hàng"', "Chuyển đến Dashboard quản lý đơn hàng (M1)"),
            ("3", 'Click "Xem danh sách đơn hàng"', "Hiển thị màn hình danh sách (M2) với bảng, tìm kiếm, bộ lọc"),
            ("4", 'Click "Xem" tại dòng DH-2026-003', "Mở modal Chi tiết đơn hàng (M3)"),
            ("5", "Quan sát thông tin chung", "Mã đơn DH-2026-003, Trạng thái Đã xử lý, Site S02, PT Tàu, ngày tạo/gửi"),
            ("6", "Quan sát bảng mặt hàng", "Hiển thị MH004 — Màn hình LCD 16x2, SL 50, Đơn vị Cái"),
        ],
        "post": "Chi tiết đơn hợp lệ được hiển thị; không dữ liệu nào bị thay đổi.",
    },
    {
        "id": "UC_TC02",
        "title": "Tìm kiếm không thấy đơn hàng",
        "pre": "Đang ở màn hình danh sách đơn hàng (M2).",
        "steps": [
            ("1", "Nhập từ khóa `DH-9999` vào ô Từ khóa", "Ô tìm kiếm nhận từ khóa"),
            ("2", 'Nhấn "Tìm kiếm"', "Hệ thống tra cứu, không có kết quả"),
            ("3", "Quan sát vùng bảng", 'Hiển thị "Không tìm thấy đơn hàng nào" + gợi ý (M4)'),
            ("4", 'Nhấn "Làm mới" (luồng 6a)', "Xóa bộ lọc/từ khóa, hiển thị lại 13 đơn trên M2"),
        ],
        "post": "Người dùng có thể tìm lại; không dữ liệu nào bị thay đổi.",
    },
    {
        "id": "UC_TC03",
        "title": "Xem chi tiết đơn hàng đã bị hủy",
        "pre": 'Trong danh sách có đơn DH-2026-008 trạng thái "Đã hủy".',
        "steps": [
            ("1", 'Click "Xem" tại dòng DH-2026-008', "Hệ thống phát hiện đơn đã hủy (luồng 7a)"),
            ("2", "Quan sát hộp thoại", 'Modal cảnh báo (M5): "Không thể xem chi tiết" + mã đơn DH-2026-008'),
            ("3", 'Nhấn "Xác nhận & Quay lại" (luồng 7b)', "Đóng hộp thoại, quay về M2"),
        ],
        "post": "Không mở M3; người dùng quay lại danh sách để chọn đơn khác.",
    },
    {
        "id": "UC_TC04",
        "title": "Lọc theo trạng thái rồi xem chi tiết",
        "pre": "Đang ở màn hình danh sách đơn hàng (M2).",
        "steps": [
            ("1", 'Chọn bộ lọc Trạng thái = "Đã xử lý"', "Dropdown nhận giá trị lọc"),
            ("2", 'Nhấn "Tìm kiếm"', "Bảng chỉ còn DH-2026-003 và DH-2026-011"),
            ("3", 'Click "Xem" một đơn trong kết quả', "Mở modal chi tiết (M3) đúng đơn đã chọn"),
        ],
        "post": "Chi tiết đơn được hiển thị đúng theo bộ lọc.",
    },
    {
        "id": "UC_TC05",
        "title": "Quay lại danh sách sau khi xem chi tiết",
        "pre": "Đang ở modal chi tiết (M3) của một đơn hợp lệ (ví dụ DH-2026-003).",
        "steps": [
            ("1", 'Nhấn "← Quay lại danh sách"', "Đóng modal chi tiết (M3)"),
            ("2", "Quan sát màn hình", "Trở về màn hình danh sách đơn hàng (M2)"),
        ],
        "post": "Người dùng ở lại M2; không dữ liệu nào bị thay đổi.",
    },
]


def write_summary_sheet(ws):
    ws.merge_cells("A1:E1")
    ws["A1"] = "15.4. Kiểm thử Use Case (System/Scenario Testing)"
    ws["A1"].font = TITLE

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "Use case: Xem chi tiết đơn hàng (UC006) | "
        "Tác nhân: Bộ phận đặt hàng quốc tế | "
        "Giao diện: M0–M5 (CardLayout) | Kết quả: 5/5 PASSED"
    )
    ws["A2"].font = NORMAL
    ws["A2"].alignment = WRAP

    ws.merge_cells("A3:E3")
    ws["A3"] = (
        "Ngoài Unit Test ở mức hàm (mục 15.3), UC006 được kiểm tra end-to-end theo luồng sự kiện "
        "từ góc nhìn người dùng (Sequence Diagram mục 10, Activity Diagram mục 3)."
    )
    ws["A3"].font = NORMAL
    ws["A3"].alignment = WRAP

    headers = [
        "UC Test ID",
        "Kịch bản kiểm thử (Scenario)",
        "Điều kiện ban đầu (Pre-conditions)",
        "Kết quả mong đợi (Post-conditions / Output)",
        "KQ Test",
    ]
    start = 5
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    style_table(ws, start, 1, start + len(SUMMARY_ROWS), 5, header_row=start)

    for idx, row in enumerate(SUMMARY_ROWS, 1):
        r = start + idx
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 58
    ws.column_dimensions["E"].width = 10


def write_detail_sheet(ws):
    ws.merge_cells("A1:E1")
    ws["A1"] = "Chi tiết bước kiểm thử Use Case — UC006 (tham khảo BT7)"
    ws["A1"].font = TITLE

    row = 3
    for case in DETAIL_CASES:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"{case['id']} — {case['title']}").font = BOLD
        row += 1

        ws.cell(row=row, column=1, value="Pre-conditions:").font = BOLD
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=case["pre"]).font = NORMAL
        row += 1

        for i, h in enumerate(["Bước", "Hành động", "Phản hồi mong đợi của hệ thống", "KQ Test"], 1):
            ws.cell(row=row, column=i, value=h).font = BOLD
        style_table(ws, row, 1, row, 4, header_row=row)
        row += 1

        for step, action, expected in case["steps"]:
            ws.cell(row=row, column=1, value=step)
            ws.cell(row=row, column=2, value=action)
            ws.cell(row=row, column=3, value=expected)
            ws.cell(row=row, column=4, value="Pass")
            style_table(ws, row, 1, row, 4)
            row += 1

        ws.cell(row=row, column=1, value="Post-conditions:").font = BOLD
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=case["post"]).font = NORMAL
        row += 2

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 10


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "15.4 Use Case Testing"
    write_summary_sheet(ws)

    ws2 = wb.create_sheet("Chi-tiet-tung-TC")
    write_detail_sheet(ws2)

    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
