#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate Excel tables for SRS section 7.7 - 7.13 (format khớp SRS_Final docx)."""
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / ".excel_tools"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

OUT = Path(__file__).resolve().parent / "7-DacTaManHinh-Muc7.7-7.13.xlsx"

BOLD = Font(bold=True, size=11)
NORMAL = Font(size=11)
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_border(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def write_screen_spec(ws, screen_name, image_ref, controls, created_date="05/06/2026"):
    """Đặc tả màn hình — format 2 bảng như SRS_Final (1).docx mục 7.6–7.10."""
    # --- Bảng metadata (2 hàng × 4 cột) ---
    ws.merge_cells("A1:B1")
    ws["A1"] = "Hệ thống phần mềm đặt hàng nhập khẩu"
    ws["A1"].font = NORMAL
    ws["C1"] = "Ngày tạo"
    ws["C1"].font = NORMAL
    ws["D1"] = "Người phụ trách"
    ws["D1"].font = NORMAL

    ws["A2"] = "Đặc tả màn hình"
    ws["A2"].font = NORMAL
    ws["B2"] = screen_name
    ws["B2"].font = NORMAL
    ws["C2"] = created_date
    ws["C2"].font = NORMAL
    ws["D2"] = "Nguyễn Đức Toàn"
    ws["D2"].font = NORMAL

    for r in (1, 2):
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = WRAP
    apply_border(ws, 1, 1, 2, 4)

    # --- Bảng điều khiển (6 cột: A–C ảnh gộp | D–F Điều khiển/Hành động/Chức năng) ---
    start = 4
    end = start + len(controls)  # hàng start = header; start+1..end = dữ liệu

    ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=3)
    img = ws.cell(row=start, column=1, value=f"Xem chi tiết tại mục {image_ref}")
    img.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    img.font = NORMAL

    for col, label in enumerate(["Điều khiển", "Hành động", "Chức năng"], 4):
        cell = ws.cell(row=start, column=col, value=label)
        cell.font = BOLD
        cell.alignment = CENTER

    for idx, (ctrl, action, func) in enumerate(controls, 1):
        r = start + idx
        ws.cell(row=r, column=4, value=ctrl).font = NORMAL
        ws.cell(row=r, column=5, value=action).font = NORMAL
        ws.cell(row=r, column=6, value=func).font = NORMAL
        for c in range(4, 7):
            ws.cell(row=r, column=c).alignment = WRAP

    for r in range(start, end + 1):
        for c in range(4, 7):
            ws.cell(row=r, column=c).border = BORDER
    apply_border(ws, start, 1, end, 3)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 62


def write_field_spec(ws, screen_name, rows, mapping=None):
    """Định nghĩa trường — format như SRS mục 7.11 (Tên màn hình + 5 cột)."""
    ws["A1"] = "Tên màn hình"
    ws["A1"].font = NORMAL
    ws["B1"] = screen_name
    ws["B1"].font = NORMAL
    ws.merge_cells("C1:E1")
    apply_border(ws, 1, 1, 1, 5)

    headers = ["Tên Item", "Số kí tự (bytes)", "Loại", "Thuộc tính trường", "Nhận xét"]
    start = 2
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=i, value=h)
        cell.font = BOLD
        cell.alignment = CENTER
    apply_border(ws, start, 1, start, 5)

    for idx, row in enumerate(rows, 1):
        r = start + idx
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = NORMAL
            cell.alignment = WRAP
        apply_border(ws, r, 1, r, 5)

    if mapping:
        r = start + len(rows) + 2
        ws.cell(row=r, column=1, value="Ánh xạ mã trạng thái (trong code)").font = BOLD
        r += 1
        for i, h in enumerate(["Mã hệ thống", "Nhãn hiển thị"], 1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.font = BOLD
            cell.alignment = CENTER
        apply_border(ws, r, 1, r, 2)
        for code, label in mapping:
            r += 1
            ws.cell(row=r, column=1, value=code).font = NORMAL
            ws.cell(row=r, column=2, value=label).font = NORMAL
            apply_border(ws, r, 1, r, 2)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 48


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # 7.7 — Trang chủ (M0)
    ws = wb.create_sheet("7.7-TrangChu-M0")
    write_screen_spec(ws, "Trang chủ hệ thống (giao diện chung nhóm)", "7.1", [
        ('Header "HỆ THỐNG ĐẶT HÀNG NHẬP KHẨU"', "Hiển thị", "Tiêu đề hệ thống, nền xanh (#2962FF)"),
        ('Nhãn "NV Đặt hàng quốc tế"', "Hiển thị", "Hiển thị tác nhân đang đăng nhập"),
        ('Ô "Phân bổ đơn đặt hàng"', "Disabled", "Placeholder — UC nhóm trưởng"),
        ('Ô "Tạo yêu cầu nhập hàng"', "Disabled", "Placeholder — UC khác"),
        ('Ô "Lọc Site theo mặt hàng"', "Disabled", "Placeholder — UC khác"),
        ('Ô "Nhập số lượng tồn kho"', "Disabled", "Placeholder — UC khác"),
        ('Ô "Xem chi tiết đơn hàng"', "Click", "Chuyển đến Dashboard quản lý đơn hàng (M1)"),
        ('Ô "Kiểm hàng"', "Disabled", "Placeholder — UC khác"),
    ])

    # 7.8 — Dashboard (M1)
    ws = wb.create_sheet("7.8-Dashboard-M1")
    write_screen_spec(ws, "Dashboard quản lý đơn hàng", "7.2", [
        ('Nút "← Về trang chủ"', "Click", "Quay lại Trang chủ hệ thống (M0)"),
        ('Tiêu đề "DASHBOARD QUẢN LÝ ĐƠN HÀNG"', "Hiển thị", "Tiêu đề trang, chữ in hoa màu xanh"),
        ('Card "Tổng đơn hàng"', "Hiển thị", "Số lượng tất cả đơn hàng trong hệ thống"),
        ('Card "Đã xử lý"', "Hiển thị", "Số đơn có trạng thái DA_GUI"),
        ('Card "Đang xử lý"', "Hiển thị", "Số đơn có trạng thái DANG_XU_LY"),
        ('Card "Yêu cầu mới"', "Hiển thị", "Số đơn có trạng thái NHAP (Nháp)"),
        ('Nút "Xem danh sách đơn hàng"', "Click", "Chuyển đến Màn hình danh sách đơn hàng (M2)"),
        ('Bảng "Đơn hàng gần đây"', "Hiển thị",
         "Liệt kê đơn mới nhất (Mã đơn, Tên Site, Phương tiện, Ngày tạo, Trạng thái); chỉ xem, không mở chi tiết trực tiếp"),
    ])

    # 7.9 — Danh sách (M2)
    ws = wb.create_sheet("7.9-DanhSach-M2")
    write_screen_spec(ws, "Màn hình danh sách đơn hàng", "7.3", [
        ('Nút "← Về trang chủ"', "Click", "Quay lại Trang chủ hệ thống (M0)"),
        ('Nút "← Quay lại Dashboard"', "Click", "Quay lại Dashboard quản lý đơn hàng (M1)"),
        ('Ô "Từ khóa"', "Nhập văn bản", "Nhập mã đơn hàng, mã Site hoặc từ khóa để tra cứu"),
        ("Bộ lọc Trạng thái", "Chọn (Dropdown)", "Lọc theo: Tất cả / Nháp / Đang xử lý / Đã xử lý / Đã hủy"),
        ("Bộ lọc Phương tiện VT", "Chọn (Dropdown)", "Lọc theo: Tất cả / Tàu / Hàng không"),
        ('Nút "Tìm kiếm"', "Click", "Tra cứu theo điều kiện đã nhập; hiển thị kết quả trên bảng. Nếu rỗng → M4"),
        ('Nút "Làm mới"', "Click", "Xóa từ khóa và bộ lọc, tải lại toàn bộ danh sách đơn hàng"),
        ('Nút "Xem" / Double-click dòng', "Click", "Đơn hợp lệ → M3; đơn đã hủy → M5"),
        ("Phân trang (Trước / Trang N / Sau)", "Click", "Chuyển trang danh sách (nếu có nhiều trang)"),
    ])

    # 7.10 — Chi tiết (M3)
    ws = wb.create_sheet("7.10-ChiTiet-M3")
    write_screen_spec(ws, "Màn hình chi tiết đơn hàng", "7.4", [
        ('Header "CHI TIẾT ĐƠN HÀNG"', "Hiển thị", "Thanh tiêu đề modal, nền xanh"),
        ('Nút "✕" (Đóng)', "Click", "Đóng modal, quay về M2"),
        ('Card "Thông tin chung"', "Hiển thị",
         "Mã đơn hàng, Trạng thái, Mã Site, Tên Site, Phương tiện VT, Số ngày vận chuyển, Ngày tạo đơn, Ngày gửi đơn"),
        ('Bảng "Danh sách mặt hàng"', "Hiển thị",
         "STT, Mã hàng, Tên mặt hàng, Số lượng đặt, Đơn vị, Phương tiện VT"),
        ('Nút "← Quay lại danh sách"', "Click", "Đóng modal, quay về Màn hình danh sách đơn hàng (M2)"),
    ])

    # 7.11 — Không tìm thấy (M4)
    ws = wb.create_sheet("7.11-KhongTimThay-M4")
    write_screen_spec(ws, "Trạng thái không tìm thấy đơn hàng", "7.5", [
        ("Ô tìm kiếm", "Hiển thị", "Giữ nguyên từ khóa người dùng đã nhập"),
        ("Khu vực thông báo trống", "Hiển thị",
         'Icon + dòng "Không tìm thấy đơn hàng nào" + gợi ý "Vui lòng kiểm tra lại từ khóa hoặc bộ lọc."'),
        ('Nút "Tìm kiếm" (trên toolbar)', "Click", "Cho phép nhập từ khóa mới và tra cứu lại (Luồng thay thế 6a)"),
        ('Nút "Làm mới"', "Click", "Xóa bộ lọc và từ khóa, hiển thị lại danh sách đầy đủ trên M2"),
    ])

    # 7.12 — Cảnh báo hủy (M5)
    ws = wb.create_sheet("7.12-CanhBao-M5")
    write_screen_spec(ws, "Hộp thoại cảnh báo đơn hàng đã bị hủy", "7.6", [
        ('Tiêu đề "Cảnh báo"', "Hiển thị", "Tiêu đề hộp thoại modal"),
        ('Nút "✕" (Đóng)', "Click", "Đóng hộp thoại, quay về M2"),
        ("Icon cảnh báo (viền đỏ)", "Hiển thị", "Nhấn mạnh tình trạng lỗi"),
        ('Dòng "Không thể xem chi tiết"', "Hiển thị", "Tiêu đề thông báo lỗi"),
        ("Thông báo chi tiết", "Hiển thị",
         "Đơn hàng [Mã đơn] đã bị hủy, không thể xem chi tiết. (Luồng thay thế 7a)"),
        ('Nút "Xác nhận & Quay lại"', "Click",
         "Đóng hộp thoại, quay về Màn hình danh sách (M2); không mở M3 (Luồng thay thế 7b)"),
    ])

    # 7.13 — Định nghĩa trường (format SRS)
    ws = wb.create_sheet("7.13.1-Dashboard")
    write_field_spec(ws, "Dashboard quản lý đơn hàng", [
        ("Tổng đơn hàng", 5, "Số", "Chữ lớn màu xanh", "Tổng số đơn trong hệ thống"),
        ("Đã xử lý", 5, "Số", "Chữ lớn màu xanh lá", "Trạng thái DA_GUI"),
        ("Đang xử lý", 5, "Số", "Chữ lớn màu cam", "Trạng thái DANG_XU_LY"),
        ("Yêu cầu mới", 5, "Số", "Chữ lớn màu xanh đậm", "Trạng thái NHAP (Nháp)"),
        ("Mã đơn hàng (bảng gần đây)", 20, "Chữ", "Màu đen", "Ví dụ: DH-2026-001"),
        ("Tên Site", 50, "Chữ", "Màu đen", "Tên Site cung cấp"),
        ("Phương tiện", 15, "Chữ", "Màu đen", "Tàu hoặc Hàng không"),
        ("Ngày tạo", 10, "Ngày", "Màu đen", "Định dạng dd/MM/yyyy"),
        ("Trạng thái", 15, "Chữ", "Badge màu", "Nháp / Đang xử lý / Đã xử lý / Đã hủy"),
    ])

    ws = wb.create_sheet("7.13.2-DanhSach")
    write_field_spec(ws, "Màn hình danh sách đơn hàng", [
        ("Mã đơn hàng", 20, "Chữ", "Màu đen, in đậm", 'Click "Xem" hoặc double-click để xem chi tiết'),
        ("Mã Site", 15, "Chữ", "Màu đen", "Mã Site nhập khẩu"),
        ("Tên Site", 50, "Chữ", "Màu đen", ""),
        ("Số mặt hàng", 5, "Số", "Màu đen", "Số lượng mặt hàng trong đơn"),
        ("Phương tiện VT", 15, "Chữ", "Màu đen", "Tàu hoặc Hàng không"),
        ("Ngày tạo", 10, "Ngày", "Màu đen", "Định dạng dd/MM/yyyy"),
        ("Trạng thái", 15, "Chữ", "Badge màu", "Nháp (xanh dương), Đang xử lý (cam), Đã xử lý (xanh lá), Đã hủy (đỏ)"),
        ("Thao tác", 10, "Chữ", 'Link "Xem"', "Mở M3 hoặc M5 tùy trạng thái đơn"),
        ("Ô tìm kiếm", 100, "Chữ", "Ô nhập, viền xám", "Nhập mã đơn, mã Site hoặc từ khóa"),
    ], mapping=[
        ("NHAP", "Nháp"),
        ("DANG_XU_LY", "Đang xử lý"),
        ("DA_GUI", "Đã xử lý"),
        ("DA_HUY", "Đã hủy"),
    ])

    ws = wb.create_sheet("7.13.3-ChiTiet")
    write_field_spec(ws, "Màn hình chi tiết đơn hàng", [
        ("Mã đơn hàng", 20, "Chữ", "Màu đen, in đậm", "Ví dụ: DH-2026-003"),
        ("Trạng thái", 15, "Chữ", "Badge màu", "Nháp / Đang xử lý / Đã xử lý / Đã hủy"),
        ("Mã Site", 15, "Chữ", "Màu đen", "Ví dụ: S02"),
        ("Tên Site", 50, "Chữ", "Màu đen", "Ví dụ: Site S02"),
        ("Phương tiện VT", 20, "Chữ", "Màu đen", "Tàu hoặc Hàng không"),
        ("Số ngày vận chuyển", 5, "Số", "Màu đen", "Đơn vị: ngày (tính theo Site + phương tiện)"),
        ("Ngày tạo đơn", 10, "Ngày", "Màu đen", "Định dạng dd/MM/yyyy"),
        ("Ngày gửi đơn", 10, "Ngày", "Màu đen", 'Hiển thị "—" nếu chưa gửi'),
        ("STT (bảng mặt hàng)", 3, "Số", "Màu đen", "Số thứ tự"),
        ("Mã hàng", 10, "Chữ", "Màu đen", "Ví dụ: MH004"),
        ("Tên mặt hàng", 50, "Chữ", "Màu đen", "Ví dụ: Màn hình LCD 16x2"),
        ("Số lượng đặt", 10, "Số", "Màu đen", "Số lượng đặt mua"),
        ("Đơn vị", 10, "Chữ", "Màu đen", "Ví dụ: Cái"),
        ("Phương tiện VT (mặt hàng)", 15, "Chữ", "Màu đen", "Tàu hoặc Hàng không"),
    ])

    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
